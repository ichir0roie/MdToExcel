from excelStructure import *
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook


import mdToArray
import os
import shutil

from openpyxl.styles import Font

import settings as stg


class ArrayToExcel:
    def __init__(self):
        self.book: BookMdInfo = None
        self.books: list[BookMdInfo] = []
        self.templateBookPath = None
        self.templateSheetName = None
        self.startRow = None
        self.startCol = None

        self.baseFont = None

        # self.templateWb=None
        self.__readTemplate()
        return

    def reset(self):
        self.book = None
        self.books = []

    def setBook(self, book: BookMdInfo):
        self.book = book

    def addBook(self, book: BookMdInfo):
        self.books.append(book)

    def __readTemplate(self):
        # self.wb = load_workbook(filename=path, keep_vba=True)
        self.templateBookPath = stg.templateBookPath
        self.templateSheetName = stg.templateSheetName
        self.startCol = stg.startCol
        self.startRow = stg.startRow

        # cant't copy sheet from book A to book B.
        self.templateWb = load_workbook(
            filename=self.templateBookPath, keep_vba=True, read_only=False
        )

    def generateBook(self, outputPath: str):

        outputDir = outputPath[0 : outputPath.rfind("/")]
        if not os.path.exists(outputPath):
            if not os.path.exists(outputDir):
                os.makedirs(outputDir)
            shutil.copy(self.templateBookPath, outputPath)
            # wb=load_workbook(filename=self.templateBookPath, keep_vba=True,read_only=False)

        wb = load_workbook(outputPath, keep_vba=True, read_only=False)

        for sheet in self.book.sheets:  # type:Sheet
            self.parseSheetMdInfoToExcel(wb, sheet)

        if stg.templateSheetName in wb.sheetnames:
            std = wb.get_sheet_by_name(stg.templateSheetName)
            wb.remove_sheet(std)

        beforePath = outputPath[0 : outputPath.rfind(".")]
        extension = outputPath[outputPath.rfind(".") + 1 :]
        if extension == "xlsx":
            os.remove(outputPath)
            extension = "xlsm"
        outputPath = "{}.{}".format(beforePath, extension)
        wb.save(outputPath)

        wb.close()

        return

    def parseSheetMdInfoToExcel(self, wb: Workbook, sheet: SheetMdInfo):
        print(sheet.sheetName)
        print(sheet.data)

        # delete before created sheet
        if sheet.sheetName in wb.sheetnames:
            std = wb.get_sheet_by_name(sheet.sheetName)
            wb.remove_sheet(std)

        ws = None
        if self.templateSheetName in wb.sheetnames:
            # ws = wb.copy_worksheet(wb.get_sheet_by_name(self.templateSheetName))
            ws = self.templateWb.get_sheet_by_name(stg.templateSheetName)
            wb.worksheets.append(ws)
        else:
            ws = wb.create_sheet()
        ws.title = sheet.sheetName
        # rootFont=ws.cell(self.startRow+1, self.startCol+1).font
        # self.baseFont=Font(name=rootFont.name,sz=rootFont.sz)

        self.baseFont = Font(name=stg.font, size=stg.size)

        for r, row in enumerate(sheet.data):
            for c, column in enumerate(row):
                if column != "":
                    self.__setVal(ws, r + 1, c + 1, column)

    def generateBooks(self, outputPath: str, font: str, size):

        outputDir = outputPath[0 : outputPath.rfind("/")]
        if not os.path.exists(outputPath):
            if not os.path.exists(outputDir):
                os.makedirs(outputDir)
            shutil.copy(self.templateBookPath, outputPath)
            # wb=load_workbook(filename=self.templateBookPath, keep_vba=True,read_only=False)

        wb = load_workbook(outputPath, keep_vba=True, read_only=False)
        for book in self.books:
            for sheet in book.sheets:
                self.parseSheetMdInfoToExcel(wb, sheet)

        wb.save(outputPath)

        wb.close()

        return

    def __setVal(self, ws: openpyxl.worksheet, row, col, val):
        cell = ws.cell(row=row + self.startRow, column=col + self.startCol)
        cell.font = self.baseFont
        cell.value = val
        return


if __name__ == "__main__":
    mte = mdToArray.MdToArray()
    mte.read("test/test.md")

    # import pickle

    # # with open("book.pickle","wb")as f:
    # #     pickle.dump(mte.book, f)
    # with open("book.pickle", "rb") as f:
    #     book = pickle.load(f)
    ate = ArrayToExcel()
    ate.setBook(book=mte.book)
    # ate.readTemplate()
    ate.generateBook("test/test.xlsx")
